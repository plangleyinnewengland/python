import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

EXCEL_FILE = "records.xlsx"
CATEGORIES_FILE = "categories.txt"

def load_categories():
    if os.path.exists(CATEGORIES_FILE):
        with open(CATEGORIES_FILE, "r") as f:
            categories = [line.strip() for line in f if line.strip()]
        if not categories:
            categories = ["General"]
    else:
        categories = ["General"]
    return categories

def save_categories(categories):
    with open(CATEGORIES_FILE, "w") as f:
        for cat in categories:
            f.write(cat + "\n")

def add_category():
    new_cat = simpledialog.askstring("Add Category", "Enter new category:")
    if new_cat:
        new_cat = new_cat.strip()
        if new_cat and new_cat not in category_combo['values']:
            categories = list(category_combo['values']) + [new_cat]
            category_combo['values'] = categories
            category_var.set(new_cat)
            save_categories(categories)

def save_record():
    date = date_entry.get()
    dc = dc_var.get()
    amount_str = amount_entry.get()
    category = category_var.get()

    if not date or not amount_str or not category:
        messagebox.showerror("Error", "Please fill all fields.")
        return

    try:
        amount = float(amount_str)
        # Round to two decimal places
        amount = round(amount, 2)
    except ValueError:
        messagebox.showerror("Error", "Amount must be a number.")
        return

    # If Debit, ensure amount is negative
    if dc == "Debit" and amount > 0:
        amount = -amount
    # If Credit, ensure amount is positive
    if dc == "Credit" and amount < 0:
        amount = abs(amount)

    # Create or load workbook
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Debit/Credit", "Amount", "Category"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    ws.append([date, dc, round(amount, 2), category])  # Store amount as float with 2 decimals
    wb.save(EXCEL_FILE)
    messagebox.showinfo("Saved", "Record saved successfully.")
    date_entry.delete(0, tk.END)
    amount_entry.delete(0, tk.END)

root = tk.Tk()
root.title("Record Entry Form")

tk.Label(root, text="Date (YYYY-MM-DD):").grid(row=0, column=0, padx=5, pady=5)
date_entry = tk.Entry(root)
date_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(root, text="Debit/Credit:").grid(row=1, column=0, padx=5, pady=5)
dc_var = tk.StringVar(value="Debit")
dc_combo = ttk.Combobox(root, textvariable=dc_var, values=["Debit", "Credit"], state="readonly")
dc_combo.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Amount:").grid(row=2, column=0, padx=5, pady=5)
amount_entry = tk.Entry(root)
amount_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Label(root, text="Category:").grid(row=3, column=0, padx=5, pady=5)
category_var = tk.StringVar()
categories = load_categories()
category_combo = ttk.Combobox(root, textvariable=category_var, values=categories, state="readonly")
category_combo.grid(row=3, column=1, padx=5, pady=5)
category_combo.current(0)

add_cat_btn = tk.Button(root, text="Add Category", command=add_category)
add_cat_btn.grid(row=3, column=2, padx=5, pady=5)

save_btn = tk.Button(root, text="Save", command=save_record)
save_btn.grid(row=4, column=0, columnspan=3, pady=10)

root.mainloop()